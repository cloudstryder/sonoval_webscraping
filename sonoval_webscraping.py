"""Imports"""
# to read + write excel data
from openpyxl.workbook import workbook
from openpyxl import load_workbook 
# to handle link redirects
from selenium import webdriver
# time.sleep()
import time 
# to webscrape
import re 
# to load websites
from requests_html import HTMLSession

# setting up notebook + HTML loader
workbook = load_workbook(filename='excel_files/research_papers_list.xlsx') 
sheet = workbook.active  
session = HTMLSession() 
# index to search for emails. it still picks up some data tags or long strings starting with // which i should filter out the next time
EMAIL_REGEX = r"""(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"""

# loads webdriver, needs to be installed, and the final breadcrumb of the filepath should be the actual webdriver
browser = webdriver.Chrome(executable_path='/Users/Rohin/Desktop/sonoval_webscraping/chromedriver/chromedriver')

# the data starts at row 2, theres probably some better way to get the value of the current index with generators but wasn't sure
# this value is used to determine where to insert the extracted emails
iterator = 2
# use generator instead of for loop to save processing power
for row in sheet.iter_rows(min_row=iterator, max_row=500, min_col=7,max_col=7):
    emails = [] # emails extracted from website
    
    """because doi.org redirects to another website, i had initially tried to use the .url 
    property of request.get objects, but they had a countermeasure against this so I had to 
    use a webdriver which did add some extra time + another library"""
    # website = requests.get("http://doi.org/"+row[0].value, verify=True)
    # URL = website.url

    """launches website in chrome window, and after the sleep time, takes the final url after the redirect"""
    print("fetching redirect from http://doi.org/"+row[0].value)
    browser.get("http://doi.org/"+row[0].value)
    time.sleep(1.5) # time waiting for browser to load all of the redirects
    URL = browser.current_url
    print("Final URL:", URL)
    
    # starts new session with final URL
    site = session.get(URL)
    # specific sites gave errors, especially when i was testing, but im not sure if its too necessary
    try:
        print("rendering site...")
        time.sleep(1.75) # for some reason you need to have a delay before you render another website otherwise it gives an error
        site.html.render()
        print("site rendered")
        # scrapes emails, and returns as array
        for re_match in re.finditer(EMAIL_REGEX, site.html.raw_html.decode()): # throws error if element w given ID is not there (i.e. email)
            if not ['/', '\\', '{','}','[',']'] in re_match.group() and not re_match.group()[-1] in ['0','1','2','3','4','5','6','7','8','9']:
                emails.append(re_match.group())
        if len(emails) == 0:
            print("Emails hidden") # if an error is not thrown but no emails are found, this usually means they are hidden by a captcha or something
            sheet["C"+str(iterator)] = "Emails hidden (CAPTCHA or other anti-webscraping measure)"
        else:    
            print(", ".join(set(emails)))
            sheet["C"+str(iterator)] = ", ".join(set(emails)) # puts emails into excel document on correct line

    except Exception as error:
        print("EXCEPTION OCCURRED:\n", error)
        sheet["C"+str(iterator)] = "ERROR"
        pass
    iterator += 1
    print('[website completed]')
    
# browser must be quit after all emails are scraped, as it cannot be reopened in one file ig
browser.quit()
# saves changes to new file (which has already been created), I could have saved this to the original file but decided to make a new one
workbook.save(filename='research_papers_list_updated.xlsx')
    


# other random code

# sheet.cell(row=r, column=c).value
# sheet["A1"].value

# for row in sheet.iter_rows(min_row=1, max_row=100, min_col=7,max_col=7):
#     if row[0].value != "DI":
#         print("http://doi.org/" + row[0].value) 



# for i in range(1, 21):
#     print(sheet.cell(row=i, column=7).value)


# URL = "https://www.google.com/search?q=shoebill&sxsrf=ALeKk02ZsvBSE9LjGQ4PSlbHAiVLKlPH7g%3A1624570319455&ei=z_nUYLSbG6Cu0PEPu4S2yA4&oq=shoebill&gs_lcp=Cgdnd3Mtd2l6EAMyBwgjELADECcyBwgjELADECcyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEENKBQg4EgExSgQIQRgAUABYAGDqxX9oAXACeACAAbsFiAG7BZIBAzUtMZgBAKoBB2d3cy13aXrIAQ_AAQE&sclient=gws-wiz&ved=0ahUKEwj0vraGnLHxAhUgFzQIHTuCDekQ4dUDCA4&uact=5"

# page = requests.get(URL)
# soup = BeautifulSoup(page.content, 'html.parser')
# print(soup.prettify())

# results = soup.find(id="center_col")
# print(results)

# #search_elems = results.find_all('div', class_="TzHB6b cLjAic")

# #print(search_elems[0])