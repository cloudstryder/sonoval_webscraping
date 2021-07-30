"""
Notes:
Although many of these comments might be a bit redundant and over-explanatory
for someone with a lot of coding experience, I would still recommend reading 
all of them throroughly before starting to be safe. 

The easiest way to run this is to just use "python3 thisfilename" in terminal. 
If any modules are missing when you first run the code, just type "pip install 
modulename". And before that, be sure to change any of the variables below as 
needed to match your local filenames/paths. 

All of this was developed on IOS, so some of the code for filepaths might 
have to be altered slightly if you're on windows or linux (it might actually
work fine with linux, I'm not too certain). 

After the code completes, if you try opening whichever file the emails were 
written to it'll likely say that the file was damaged, but if you open and
delete the recovered data it should be fine (I usually check the recovered 
just in case theres something important there, but its usually just some 
random text which opens in xcode and can be discarded).  

After downloading the code, if you want to run everything while changing as 
little as possible, make a new directory here called "excel_files" and name
the excel file "research_papers_list.xlsx". The new file will then be saved
as "research_papers_list_updated.xlsx".

If the code gets interrupted, do not start again and save the data to the 
same file. This will override all of the previous data which was entered 
before the interruption with whatever is on the initial file (so usually
empty cells). The solution would be to change the read file to the file 
made by the previously interrupted code and the write file to a new one.
"""


"""variables"""
# file path for excel file which has all of the DOI #s for the articles
# (note that all excel files end with .xlsx)
READ_FILE = 'excel_files/research_papers_list.xlsx'
# file path for excel file which the emails will be copied onto. if you want the 
# data saved to a new file, make this a file path to an excel doc which doesn't exist, 
# and all of the contents of the READ_FILE will be copied over into the new file.
WRITE_FILE = 'research_papers_list_updated.xlsx'
# file path to chromeddriver module. It should be noted that it must include the 
# module itself as the final item in the path, and it must be an absolute file 
# path. You can experiment with a relative file path but the chromedriver was
# always very finnicky for me so its probably safest to just do this. 
WEBDRIVER_PATH = '/Users/Rohin/Desktop/sonoval_webscraping/chromedriver/chromedriver'
# Which row the algorithm should start reading at (usually row 2 since the first row is headers).
# This variable is also used to determine which row the emails should start being written at.
READ_ROW = 2
# Row at which code stops reading. The final line is usually 500, but you can reduce this if you are
# testing any new code you added and want to see if file saves/webscraping part works
END_ROW = 500
# which column algorithm should start reading at (this should be a number, not a letter). 
# If the same structure is kept for the excel file, it should be at column 7
READ_COLUMN = 7
# column in which emails are written to in final file. Must be a letter, usually column C
WRITE_COLUMN = 'C'
# If any of these characters are found in one of the strings presumed to be an email, it
# will be discarded. These can also be full strings and not just chars. 
FORBIDDEN_CHARS = ['/', '\\', '{','}','[',']']
# emails ending with any of these strings will be ignored. The strings can be any length. 
FORBIDDEN_ENDINGS = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
# emails starting with any of these strings will be ignored, can be any length
FORBIDDEN_STARTERS = ["ampaign", "info", "journal"]
# if an email excedes this length, it will be discarded. 
# abcdefghijklmnopqrstuvwxyzabcdefghijklmn
# The line above is what 40 characters looks like, but if you feel that people might have very
# long emails/domains you are welcome to change this to something unobstructive (100) to be safe
MAX_LENGTH = 40
# set to true if you would like to have the emails be saved as the code is running or for 
# everything to be saved at the very end after everything is run. It is safer to set this
# as true, especially if there will likely be an interruption in the code
SAVE_EVERY_ITERATION = True



"""optional variables"""




"""Imports"""
# to read + write excel data
from openpyxl.workbook import workbook
from openpyxl import load_workbook 
# webdriver to handle link redirects
from selenium import webdriver
# time.sleep()
import time 
# to webscrape
import re 
# to load websites
from requests_html import HTMLSession

# setting up notebook + HTML loader
workbook = load_workbook(filename=READ_FILE) 
if not SAVE_EVERY_ITERATION:
    sheet = workbook.active  

session = HTMLSession() 
# index to search for emails. it still picks up some data tags or long strings starting with // which i should filter out the next time
EMAIL_REGEX = r"""(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"""

# loads webdriver, needs to be installed, and the final breadcrumb of the filepath should be the actual webdriver
browser = webdriver.Chrome(executable_path=WEBDRIVER_PATH)

# use generator instead of for loop to save processing power
if SAVE_EVERY_ITERATION:
    sheet = workbook.active
for row in sheet.iter_rows(min_row=READ_ROW, max_row=END_ROW, min_col=READ_COLUMN,max_col=READ_COLUMN):
    emails = [] # emails extracted from website
    
    print("line", READ_ROW, "out of", END_ROW)

    """because doi.org redirects to another website, i had initially tried to use the .url 
    property of request.get objects, but they had a countermeasure against this so I had to 
    use a webdriver"""
    # website = requests.get("http://doi.org/"+row[0].value, verify=True)
    # URL = website.url

    """launches website in chrome window, and after the sleep time, takes the final url after the redirect"""
    print("fetching redirect from http://doi.org/"+row[0].value)
    browser.get("http://doi.org/"+row[0].value)
    # time.sleep(1.5) this was not actually needed, the next line waits for the website to be fully loaded before continuing and doesnt need a timer
    URL = browser.current_url
    print("Final URL:", URL)
    
    # starts new session with final URL
    site = session.get(URL)
    # specific sites gave errors, especially when i was testing, but im not sure if this try statement is too necessary
    try:
        print("rendering site...")
        time.sleep(1.75) # for some reason you need to have a delay before you render another website otherwise it gives an error
        site.html.render()
        print("site rendered")
        # scrapes emails, and returns as array
        for re_match in re.finditer(EMAIL_REGEX, site.html.raw_html.decode()): # throws error if element w given ID is not there (i.e. email)
            if not any(x in re_match.group() for x in FORBIDDEN_CHARS) and not any(re_match.group().endswith(x) for x in FORBIDDEN_ENDINGS) and not any(re_match.group().startswith(x) for x in FORBIDDEN_STARTERS) and len(re_match.group()) < MAX_LENGTH:
                emails.append(re_match.group())
        if len(emails) == 0:
            print("Emails hidden/ignored") # if an error is not thrown but no emails are found, this usually means they are hidden by a captcha or something
            sheet[WRITE_COLUMN+str(READ_ROW)] = "Emails hidden/ignored (check website)"
        else:    
            print(", ".join(set(emails)))
            sheet[WRITE_COLUMN+str(READ_ROW)] = ", ".join(set(emails)) # puts emails into excel document on correct line
    except Exception as error: 
        print("EXCEPTION OCCURRED:\n", error)
        sheet[WRITE_COLUMN+str(READ_ROW)] = "EXCEPTION OCCURRED"
        pass
    READ_ROW += 1
    if SAVE_EVERY_ITERATION:
        workbook.save(filename=WRITE_FILE)
    print('[website completed]\n')
    
# browser must be quit after all emails are scraped, as quitting and reopening it multiple times in a 
# file gives an error. 
browser.quit()
# saves changes to file, whether it be the same one or a new one
if not SAVE_EVERY_ITERATION:
    workbook.save(filename=WRITE_FILE)
    


"""other random code which I was testing out/using as a reference"""

# sheet.cell(row=r, column=c).value
# sheet["A1"].value

# for row in sheet.iter_rows(min_row=1, max_row=100, min_col=7,max_col=7):
#     if row[0].value != "DI":
#         print("http://doi.org/" + row[0].value) 



# for i in range(1, 21):
#     print(sheet.cell(row=i, column=7).value)


# URL = "https://www.google.com/search?q=shoebill&sxsrf=ALeKk02ZsvBSE9LjGQ4PSlbHAiVLKlPH7g%3A1624570319455&ei=z_nUYLSbG6Cu0PEPu4S2yA4&oq=shoebill&gs_lcp=Cgdnd3Mtd2l6EAMyBwgjELADECcyBwgjELADECcyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyBwgAEEcQsAMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEEMyCgguELADEMgDEENKBQg4EgExSgQIQRgAUABYAGDqxX9oAXACeACAAbsFiAG7BZIBAzUtMZgBAKoBB2d3cy13aXrIAQ_AAQE&sclient=gws-wiz&ved=0ahUKEwj0vraGnLHxAhUgFzQIHTuCDekQ4dUDCA4&uact=5"

# page = requests.get(URL)
# soup = BeautifulSoup(page.content, 'html.parser')
# print(soup.prettify())

# results = soup.find(id="center_col")
# print(results)

# #search_elems = results.find_all('div', class_="TzHB6b cLjAic")

# #print(search_elems[0])